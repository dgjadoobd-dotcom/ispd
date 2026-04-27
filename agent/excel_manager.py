# ============================================================
# Excel Manager — auto-update billing reports
# ============================================================
import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_FILE, EXCEL_BACKUP

# Colors
COLOR_HEADER  = "1E3A5F"
COLOR_PAID    = "D4EDDA"
COLOR_UNPAID  = "F8D7DA"
COLOR_PARTIAL = "FFF3CD"
COLOR_WHITE   = "FFFFFF"

def _border():
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_style(cell, text):
    cell.value = text
    cell.font  = Font(bold=True, color="FFFFFF", size=11)
    cell.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()

def backup_excel():
    if not os.path.exists(EXCEL_FILE):
        return
    os.makedirs(EXCEL_BACKUP, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(EXCEL_BACKUP, f"billing_{ts}.xlsx")
    shutil.copy2(EXCEL_FILE, dest)
    print(f"📁 Backup saved: {dest}")

def update_billing_report(customers: list, stats: dict):
    """Write full billing report to Excel"""
    os.makedirs(os.path.dirname(EXCEL_FILE) or '.', exist_ok=True)
    backup_excel()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 20

    ws_sum['A1'] = "FCNCHBD ISP ERP — Daily Report"
    ws_sum['A1'].font = Font(bold=True, size=16, color=COLOR_HEADER)
    ws_sum.merge_cells('A1:B1')

    ws_sum['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws_sum['A2'].font = Font(italic=True, color="888888")
    ws_sum.merge_cells('A2:B2')

    summary_data = [
        ("Total Customers",    stats.get('total_customers', 0)),
        ("Active",             stats.get('active', 0)),
        ("Suspended",          stats.get('suspended', 0)),
        ("Today's Collection", f"৳{stats.get('today_collection', 0):,.2f}"),
        ("Month Collection",   f"৳{stats.get('month_collection', 0):,.2f}"),
        ("Total Due",          f"৳{stats.get('total_due', 0):,.2f}"),
    ]
    for i, (label, value) in enumerate(summary_data, start=4):
        ws_sum[f'A{i}'] = label
        ws_sum[f'B{i}'] = value
        ws_sum[f'A{i}'].font = Font(bold=True)
        ws_sum[f'B{i}'].alignment = Alignment(horizontal="right")
        ws_sum[f'A{i}'].border = _border()
        ws_sum[f'B{i}'].border = _border()

    # ── Sheet 2: All Customers ────────────────────────────────
    ws = wb.create_sheet("Customers")
    headers = ["#", "Customer Code", "Full Name", "Phone", "Package",
               "Monthly Charge", "Due Amount", "Status", "Updated"]
    col_widths = [5, 15, 25, 15, 20, 16, 16, 12, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _header_style(ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for i, c in enumerate(customers, start=2):
        status    = c.get('status', 'unknown')
        due       = float(c.get('due_amount', 0))
        fill_color = COLOR_PAID if status == 'active' and due == 0 else \
                     COLOR_UNPAID if due > 0 else COLOR_WHITE

        row_data = [
            i - 1,
            c.get('customer_code', ''),
            c.get('full_name', ''),
            c.get('phone', ''),
            c.get('package_name', ''),
            float(c.get('monthly_charge', 0)),
            due,
            status.upper(),
            datetime.now().strftime('%Y-%m-%d %H:%M'),
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _border()
            cell.fill   = PatternFill("solid", fgColor=fill_color)
            if col in (6, 7):  # money columns
                cell.number_format = '৳#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # ── Sheet 3: Due Customers ────────────────────────────────
    ws_due = wb.create_sheet("Due Customers")
    due_headers = ["#", "Customer Code", "Name", "Phone", "Due Amount", "Package", "Status"]
    due_widths  = [5, 15, 25, 15, 16, 20, 12]

    for col, (h, w) in enumerate(zip(due_headers, due_widths), start=1):
        _header_style(ws_due.cell(row=1, column=col), h)
        ws_due.column_dimensions[get_column_letter(col)].width = w

    ws_due.freeze_panes = 'A2'
    due_customers = [c for c in customers if float(c.get('due_amount', 0)) > 0]
    due_customers.sort(key=lambda x: float(x.get('due_amount', 0)), reverse=True)

    for i, c in enumerate(due_customers, start=2):
        row_data = [
            i - 1,
            c.get('customer_code', ''),
            c.get('full_name', ''),
            c.get('phone', ''),
            float(c.get('due_amount', 0)),
            c.get('package_name', ''),
            c.get('status', '').upper(),
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws_due.cell(row=i, column=col, value=val)
            cell.border = _border()
            cell.fill   = PatternFill("solid", fgColor=COLOR_UNPAID)
            if col == 5:
                cell.number_format = '৳#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    wb.save(EXCEL_FILE)
    print(f"✅ Excel updated: {EXCEL_FILE}")
    print(f"   → {len(customers)} customers | {len(due_customers)} with dues")
    return EXCEL_FILE
